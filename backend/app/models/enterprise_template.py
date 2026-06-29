"""Generate enterprise sheet content for intake templates."""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

ENTERPRISE_SHEET_TITLE = "企业基础资料"

ENTERPRISE_FIELDS: list[tuple[str, str, str]] = [
    ("企业名称*", "name", "必填，用于资料页眉与落款"),
    ("企业简称", "short_name", ""),
    ("法定代表人", "legal_rep", ""),
    ("联系人", "contact", ""),
    ("联系电话", "phone", ""),
    ("企业地址", "address", ""),
    ("审核年度", "year", "如 2025，默认当前年度"),
]


def write_enterprise_sheet(ws: Worksheet) -> None:
    """Populate a worksheet with the enterprise key-value template."""
    note_font = Font(name="微软雅黑", size=9, italic=True, color="7F7F7F")
    header_font = Font(bold=True, name="微软雅黑", size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    field_font = Font(name="微软雅黑", size=10)
    example_font = Font(name="微软雅黑", size=10, color="808080")

    ws.merge_cells("A1:B1")
    ws["A1"] = "请在本表填写企业基础资料，带 * 为必填项。"
    ws["A1"].font = note_font
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 36

    for col_idx, header in enumerate(["字段", "值"], start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    examples = {
        "name": "浙江精创汽车零部件有限公司",
        "short_name": "精创汽配",
        "legal_rep": "陈建国",
        "contact": "刘芳",
        "phone": "0571-88776655",
        "address": "浙江省杭州市余杭区良渚街道工矿路88号",
        "year": "2025",
    }

    for row_idx, (label, key, note) in enumerate(ENTERPRISE_FIELDS, start=3):
        field_cell = ws.cell(row=row_idx, column=1, value=label)
        field_cell.font = field_font
        if label.endswith("*"):
            field_cell.fill = PatternFill("solid", fgColor="C6EFCE")

        value_cell = ws.cell(row=row_idx, column=2, value=examples.get(key, ""))
        value_cell.font = example_font
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")

        if note:
            ws.cell(row=row_idx, column=3, value=note).font = note_font

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 28
    ws.row_dimensions[2].height = 22


def generate_enterprise_template() -> bytes:
    """Return a single-sheet enterprise template (legacy)."""
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ENTERPRISE_SHEET_TITLE
    write_enterprise_sheet(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
