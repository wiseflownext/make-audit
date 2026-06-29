"""Generate roster sheet content for intake templates."""
from __future__ import annotations

from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

ROSTER_SHEET_TITLE = "人员花名册"

ROSTER_COLUMNS: list[tuple[str, str]] = [
    ("姓名*", "name"),
    ("部门*", "department_name"),
    ("岗位*", "position_name"),
    ("岗位类别", "position_category"),
    ("入职日期*", "hire_date"),
    ("性别", "gender"),
    ("身份证号", "id_no"),
    ("联系电话", "phone"),
    ("学历", "education"),
    ("毕业院校", "school"),
    ("家庭住址", "address"),
    ("是否重点岗位", "is_key_position"),
    ("是否内审员", "is_internal_auditor"),
    ("是否普通员工", "is_regular_employee"),
    ("是否管理人员", "is_manager"),
    ("在职状态", "employment_status"),
    ("备注", "remark"),
]

NOTES_ROW: dict[str, str] = {
    "姓名*": "必填",
    "部门*": "必填，如：生产部、品质部、技术部、综合管理部",
    "岗位*": "必填，如：操作工、检验员、工艺工程师、总经理",
    "岗位类别": "生产/品质/技术/管理，用于分类统计与培训匹配",
    "入职日期*": "必填，格式 YYYY-MM-DD 或 YYYY/M/D",
    "性别": "男/女（可从身份证推导）",
    "身份证号": "18位，填写后可自动推导出生年月和性别",
    "是否重点岗位": "是/否，默认否",
    "是否内审员": "是/否，默认否",
    "是否普通员工": "是/否，默认是（普通员工试用期7天，其他30天）",
    "是否管理人员": "是/否，默认否",
    "在职状态": "在职/离职，默认在职",
}


def write_roster_sheet(ws: Worksheet) -> None:
    """Populate a worksheet with the roster column template."""
    header_font = Font(bold=True, name="微软雅黑", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    note_font = Font(name="微软雅黑", size=9, italic=True, color="7F7F7F")

    for col_idx, (header, _) in enumerate(ROSTER_COLUMNS, start=1):
        note = NOTES_ROW.get(header, "")
        cell = ws.cell(row=1, column=col_idx, value=note)
        cell.font = note_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, (header, _) in enumerate(ROSTER_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, name="微软雅黑", size=11, color="FFFFFF")
        cell.fill = header_fill if not header.endswith("*") else PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    col_widths = [12, 14, 16, 12, 14, 8, 22, 14, 10, 16, 24, 12, 10, 12, 12, 10, 16]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 22

    from app.models.sample_roster_data import get_sample_roster_rows

    examples = get_sample_roster_rows()
    for row_idx, row_data in enumerate(examples, start=3):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="微软雅黑", size=10, color="808080")
            cell.alignment = Alignment(horizontal="left", vertical="center")


def generate_roster_template() -> bytes:
    """Return a single-sheet roster template (legacy)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ROSTER_SHEET_TITLE
    write_roster_sheet(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
