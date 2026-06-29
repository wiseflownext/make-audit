"""Tests for enterprise Excel parsing."""
from __future__ import annotations

from io import BytesIO

import openpyxl
import pytest

from app.models.enterprise import EnterpriseValidationError, parse_enterprise
from app.models.enterprise_template import generate_enterprise_template
from app.models.intake_template import generate_intake_template


def _make_key_value_xlsx(rows: list[tuple[str, str]], include_header: bool = True) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    start = 1
    if include_header:
        ws.cell(row=1, column=1, value="字段")
        ws.cell(row=1, column=2, value="值")
        start = 2
    for idx, (field, value) in enumerate(rows, start=start):
        ws.cell(row=idx, column=1, value=field)
        ws.cell(row=idx, column=2, value=value)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_enterprise_success():
    content = _make_key_value_xlsx([
        ("企业名称*", "测试企业有限公司"),
        ("企业简称", "测试企业"),
        ("法定代表人", "王五"),
        ("审核年度", "2025"),
    ])
    result = parse_enterprise(content, filename="enterprise.xlsx")
    assert result["name"] == "测试企业有限公司"
    assert result["short_name"] == "测试企业"
    assert result["legal_rep"] == "王五"
    assert result["year"] == "2025"


def test_parse_enterprise_missing_name():
    content = _make_key_value_xlsx([("企业简称", "测试企业")])
    with pytest.raises(EnterpriseValidationError, match="企业名称"):
        parse_enterprise(content, filename="enterprise.xlsx")

def test_parse_generated_template():
    content = generate_intake_template()
    result = parse_enterprise(content, filename="template.xlsx", sheet_name="企业基础资料")
    assert result["name"] == "浙江精创汽车零部件有限公司"
    assert result["year"] == "2025"


def test_parse_enterprise_ignores_unknown_fields():
    content = _make_key_value_xlsx([
        ("企业名称*", "测试企业有限公司"),
        ("未知字段", "忽略"),
    ])
    result = parse_enterprise(content, filename="enterprise.xlsx")
    assert result["name"] == "测试企业有限公司"
    assert "未知字段" not in result
